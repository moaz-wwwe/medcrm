import openpyxl, io

def test_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["", "", ""])
    ws.append(["Mohamed", "01001234567", "Hospital"])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    wb2 = openpyxl.load_workbook(stream, data_only=True)
    sheet = wb2.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    data_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):
            data_rows.append(dict(zip(headers, row)))
    
    print(data_rows)

test_excel()
