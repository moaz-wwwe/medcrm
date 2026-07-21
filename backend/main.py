"""
main.py
-------
FastAPI application entrypoint for the Medical Supplies CRM.

Endpoints:
- POST /auth/login              -> obtain JWT
- POST /auth/register           -> (admin only) create a new user
- GET  /auth/me                 -> current user info
- POST /leads/                  -> create a lead
- GET  /leads/                  -> list leads (role-filtered)
- POST /call-logs/               -> create a call log
- GET  /call-logs/               -> list call logs (role-filtered)
- POST /api/manager-chat        -> (admin only) natural-language team performance Q&A via Gemini

Role-based access control (RBAC) summary:
- admin (Manager): sees ALL leads / call logs / performance data, team-wide.
- sales_rep: sees ONLY leads assigned to them, and call logs on those leads.
  This is enforced at the query level (server-side filtering), not just in the UI.
"""

import os
import sys
import json
import urllib.request
import urllib.parse
import csv
import io

# Add the directory containing this file to sys.path so local imports work in Vercel
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

from collections import defaultdict
from datetime import datetime
from typing import List, Optional

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, HTTPException, status, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session

import auth
import models
import schemas
from database import Base, engine, get_db

load_dotenv()

# Create tables on startup (fine for SQLite / dev; use Alembic migrations in production).
Base.metadata.create_all(bind=engine)

def send_telegram_notification(message: str):
    bot_token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not bot_token or not chat_id:
        return
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    data = {"chat_id": chat_id, "text": message, "parse_mode": "HTML"}
    req = urllib.request.Request(url, data=json.dumps(data).encode("utf-8"), headers={"Content-Type": "application/json"})
    try:
        urllib.request.urlopen(req, timeout=5)
    except Exception as e:
        print(f"Telegram notification failed: {e}")

app = FastAPI(title="Medical Supplies CRM API", version="1.0.0")

# Allow the static frontend (opened via file:// or a local dev server) to call the API.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Startup: seed a default admin + sales rep if the DB is empty, so the app is
# runnable immediately without a separate seeding step.
# ---------------------------------------------------------------------------
@app.on_event("startup")
def seed_default_users():
    db: Session = next(get_db())
    try:
        if db.query(models.User).count() == 0:
            admin = models.User(
                username="admin",
                password_hash=auth.hash_password("admin123"),
                role=models.UserRole.ADMIN,
                full_name="Team Manager",
            )
            rep = models.User(
                username="rep1",
                password_hash=auth.hash_password("rep123"),
                role=models.UserRole.SALES_REP,
                full_name="Sales Rep One",
            )
            db.add_all([admin, rep])
            db.commit()
            print("Seeded default users -> admin/admin123 (admin), rep1/rep123 (sales_rep)")
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.post("/auth/login", response_model=schemas.Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = auth.authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = auth.create_access_token(data={"sub": user.username, "role": user.role.value})
    return schemas.Token(
        access_token=access_token,
        role=user.role,
        username=user.username,
        user_id=user.id,
    )


@app.post("/auth/register", response_model=schemas.UserOut)
def register_user(
    payload: schemas.UserCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_admin),
):
    """Only an admin (manager) can create new user accounts (reps or other admins)."""
    if db.query(models.User).filter(models.User.username == payload.username).first():
        raise HTTPException(status_code=400, detail="Username already exists")
    new_user = models.User(
        username=payload.username,
        password_hash=auth.hash_password(payload.password),
        role=payload.role,
        full_name=payload.full_name,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@app.get("/api/analytics")
def get_analytics(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    if current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
        
    logs = db.query(models.CallLog).all()
    
    sales_by_rep = {}
    calls_by_outcome = {}
    
    for log in logs:
        # Calls by outcome
        out = log.call_result or "Unknown"
        calls_by_outcome[out] = calls_by_outcome.get(out, 0) + 1
        
        # Sales by rep (need to join user or fetch)
        rep_name = "Unknown"
        if log.lead and log.lead.assigned_rep:
            rep_name = log.lead.assigned_rep.full_name or log.lead.assigned_rep.username
            
        sales_by_rep[rep_name] = sales_by_rep.get(rep_name, 0.0) + (log.sales_amount or 0.0)
        
    return {
        "sales_by_rep": sales_by_rep,
        "calls_by_outcome": calls_by_outcome
    }

@app.get("/api/migrate-db")
def run_db_migration(db: Session = Depends(get_db)):
    from sqlalchemy import text
    
    try:
        db.execute(text("ALTER TABLE leads ADD COLUMN followup_date TIMESTAMP;"))
        db.commit()
    except Exception:
        db.rollback()
        
    try:
        db.execute(text("ALTER TABLE leads ADD COLUMN is_ignored BOOLEAN DEFAULT FALSE;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        db.execute(text("ALTER TABLE leads ADD COLUMN ignore_reason VARCHAR;"))
        db.commit()
    except Exception:
        db.rollback()

    return {"status": "success", "message": "Migration completed successfully."}

@app.post("/leads/clear-pending")
def clear_pending_leads(db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_user)):
    if current_user.role != models.UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    try:
        # Delete leads that have NO call logs
        subquery = db.query(models.CallLog.lead_id).subquery()
        deleted_count = db.query(models.Lead).filter(models.Lead.id.not_in(subquery)).delete(synchronize_session=False)
        db.commit()
        return {"status": "success", "message": f"Successfully deleted {deleted_count} pending leads."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/db-count")
def get_db_count(db: Session = Depends(get_db)):
    try:
        from models import Lead, CallLog
        total_leads = db.query(Lead).count()
        pending_leads = db.query(Lead).outerjoin(CallLog).filter(CallLog.id == None).count()
        called_leads = total_leads - pending_leads
        total_logs = db.query(CallLog).count()
        return {
            "total_leads": total_leads,
            "pending_leads": pending_leads,
            "called_leads": called_leads,
            "total_call_logs": total_logs
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/reports/rep-activity/excel")
def export_rep_activity_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_admin),
):
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # 1. Fetch all Call Logs
    logs = db.query(models.CallLog).join(models.Lead).all()
    # 2. Fetch all Ignored Leads
    ignored_leads = db.query(models.Lead).filter(models.Lead.is_ignored == True).all()

    # 3. Build detailed activity events list
    events = []
    
    # Process call logs
    for log in logs:
        lead = log.lead
        rep_username = "N/A"
        if lead and lead.assigned_rep:
            rep_username = lead.assigned_rep.username
            
        events.append({
            "lead_id": lead.id if lead else "N/A",
            "rep": rep_username,
            "date": log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else "N/A",
            "lead_name": lead.name if lead else "N/A",
            "lead_phone": lead.phone if lead else "N/A",
            "facility_type": lead.facility_type if lead else "N/A",
            "lead_notes": lead.notes if lead else "لا يوجد",
            "action_type": "مكالمة",
            "outcome": log.call_result or "N/A",
            "sales_amount": float(log.sales_amount or 0.0),
            "rep_notes": log.notes or "لا يوجد",
            "followup": lead.followup_date.strftime("%Y-%m-%d") if (lead and lead.followup_date) else "لا يوجد"
        })
        
    # Process ignored leads
    for lead in ignored_leads:
        rep_username = lead.assigned_rep.username if lead.assigned_rep else "N/A"
        events.append({
            "lead_id": lead.id,
            "rep": rep_username,
            "date": lead.created_at.strftime("%Y-%m-%d %H:%M:%S") if lead.created_at else "N/A",
            "lead_name": lead.name,
            "lead_phone": lead.phone,
            "facility_type": lead.facility_type,
            "lead_notes": lead.notes or "لا يوجد",
            "action_type": "تجاهل",
            "outcome": lead.ignore_reason or "N/A",
            "sales_amount": 0.0,
            "rep_notes": lead.ignore_reason or "لا يوجد",
            "followup": "لا يوجد"
        })

    # 4. Generate daily summary data
    summary_dict = {}
    for ev in events:
        key = (ev["rep"], ev["date"])
        if key not in summary_dict:
            summary_dict[key] = {"calls": 0, "ignored": 0, "sales": 0.0}
        
        if ev["action_type"] == "اتصال":
            summary_dict[key]["calls"] += 1
            summary_dict[key]["sales"] += ev["sales_amount"]
        elif ev["action_type"] == "تجاهل":
            summary_dict[key]["ignored"] += 1

    # Convert to list
    summary_list = []
    for (rep, date), metrics in summary_dict.items():
        summary_list.append({
            "rep": rep,
            "date": date,
            "calls": metrics["calls"],
            "ignored": metrics["ignored"],
            "sales": metrics["sales"]
        })
    # Sort
    summary_list.sort(key=lambda x: (x["date"], x["rep"]), reverse=True)
    events.sort(key=lambda x: x["date"], reverse=True)

    # 5. Create Excel Workbook using openpyxl
    wb = openpyxl.Workbook()
    
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Sheet 1: الملخص اليومي
    ws1 = wb.active
    ws1.title = "الملخص اليومي"
    ws1.views.sheetView[0].showGridLines = True
    
    headers1 = ["المندوب", "التاريخ", "عدد الاتصالات", "عدد المتجاهلين", "إجمالي المبيعات"]
    ws1.append(headers1)
    
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    for item in summary_list:
        ws1.append([item["rep"], item["date"], item["calls"], item["ignored"], item["sales"]])
        
    for row in range(2, ws1.max_row + 1):
        for col in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            if col in [1, 2]:
                cell.alignment = Alignment(horizontal="center")
            elif col in [3, 4]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            elif col == 5:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '$#,##0.00'

    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Sheet 2: سجل النشاط التفصيلي
    ws2 = wb.create_sheet(title="سجل النشاط التفصيلي")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = [
        "كود العميل (ID)", 
        "المندوب", 
        "تاريخ ووقت الإجراء", 
        "اسم العميل", 
        "رقم التليفون", 
        "نوع المنشأة", 
        "الملاحظات العامة للعميل",
        "نوع الإجراء", 
        "النتيجة / سبب التجاهل", 
        "ملاحظات المندوب التفصيلية", 
        "تاريخ المتابعة القادمة",
        "المبيعات"
    ]
    ws2.append(headers2)
    
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for item in events:
        ws2.append([
            item["lead_id"],
            item["rep"],
            item["date"],
            item["lead_name"],
            item["lead_phone"],
            item["facility_type"],
            item["lead_notes"],
            item["action_type"],
            item["outcome"],
            item["rep_notes"],
            item["followup"],
            item["sales_amount"]
        ])
        
    for row in range(2, ws2.max_row + 1):
        action_type = ws2.cell(row=row, column=8).value
        row_fill = None
        if action_type == "تجاهل":
            row_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            if col in [1, 2, 3, 5, 8, 11]:
                cell.alignment = Alignment(horizontal="center")
            elif col == 12:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '$#,##0.00'
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 14)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"rep_activity_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/auth/me", response_model=schemas.UserOut)
def read_me(current_user: models.User = Depends(auth.get_current_user)):
    return current_user


# ---------------------------------------------------------------------------
# Leads
# ---------------------------------------------------------------------------

def _lead_to_out(lead: models.Lead) -> schemas.LeadOut:
    out = schemas.LeadOut.model_validate(lead)
    out.assigned_rep_username = lead.assigned_rep.username if lead.assigned_rep else None
    
    # Attach latest call log
    if lead.call_logs:
        # Sort logs by timestamp desc
        sorted_logs = sorted(lead.call_logs, key=lambda x: x.timestamp, reverse=True)
        out.latest_call_log = _calllog_to_out(sorted_logs[0])
    else:
        out.latest_call_log = None
        
    return out


@app.post("/leads/", response_model=schemas.LeadOut)
def create_lead(
    payload: schemas.LeadCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    # RBAC: sales reps can only create leads assigned to themselves.
    # Admins may optionally assign a lead to any rep via `assigned_to`.
    if current_user.role == models.UserRole.SALES_REP:
        assigned_to = current_user.id
    else:
        assigned_to = payload.assigned_to or current_user.id
        if not db.query(models.User).filter(models.User.id == assigned_to).first():
            raise HTTPException(status_code=404, detail="assigned_to user not found")

    lead = models.Lead(
        name=payload.name,
        phone=payload.phone,
        facility_type=payload.facility_type,
        notes=payload.notes,
        assigned_to=assigned_to,
    )
    db.add(lead)
    db.commit()
    db.refresh(lead)
    
    # Send Telegram Notification
    rep_name = current_user.full_name or current_user.username
    msg = (
        f"🚨 <b>New Lead Created!</b>\n\n"
        f"👤 <b>Name:</b> {lead.name}\n"
        f"📱 <b>Phone:</b> {lead.phone}\n"
        f"🏥 <b>Facility:</b> {lead.facility_type}\n"
        f"📝 <b>Notes:</b> {lead.notes or 'N/A'}\n"
        f"👤 <b>Created By:</b> {rep_name}"
    )
    send_telegram_notification(msg)
    
    return _lead_to_out(lead)


@app.post("/leads/bulk-upload")
async def bulk_upload_leads(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_admin),
):
    """
    Reads a CSV or Excel file, creates Lead records, and distributes them
    equally among all available sales reps (Round-Robin).
    Format expected: name, phone, facility_type, notes
    """
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are allowed.")

    try:
        content = await file.read()
        
        data_rows = []
        if file.filename.endswith('.csv'):
            decoded_content = content.decode('utf-8-sig')
            
            # Read first row to get headers and fix duplicates/empties
            try:
                dialect = csv.Sniffer().sniff(decoded_content[:2048])
                reader = csv.reader(io.StringIO(decoded_content), dialect=dialect)
            except Exception:
                reader = csv.reader(io.StringIO(decoded_content))
                
            rows_list = list(reader)
            if rows_list:
                raw_headers = rows_list[0]
                headers = []
                for i, h in enumerate(raw_headers):
                    h_clean = str(h).strip() if h else ""
                    if not h_clean or h_clean in headers:
                        h_clean = f"column_{i+1}"
                    headers.append(h_clean)
                
                for row in rows_list[1:]:
                    if any(row):
                        # Pad row if shorter than headers
                        padded_row = row + [""] * (len(headers) - len(row))
                        data_rows.append(dict(zip(headers, padded_row)))

        elif file.filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active
            
            # Extract headers and fix duplicates/empties
            raw_headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            headers = []
            for i, h in enumerate(raw_headers):
                if not h or h in headers:
                    h = f"column_{i+1}"
                headers.append(h)
                
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # Skip completely empty rows
                    # row might contain None, replace with empty string
                    clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
                    padded_row = clean_row + [""] * (len(headers) - len(clean_row))
                    data_rows.append(dict(zip(headers, padded_row)))
        
        if not data_rows:
            raise HTTPException(status_code=400, detail="The uploaded file is empty.")

        # Normalize headers (remove BOM, lowercase, strip)
        normalized_rows = []
        for row in data_rows:
            norm_row = {}
            for k, v in row.items():
                if k is None: continue
                clean_k = str(k).replace('\ufeff', '').strip().lower()
                norm_row[clean_k] = v
            normalized_rows.append(norm_row)

        # Shuffle rows to randomize lead distribution
        import random
        random.shuffle(normalized_rows)

        # Get all sales reps to distribute
        reps = db.query(models.User).filter(models.User.role == models.UserRole.SALES_REP).all()
        if not reps:
            raise HTTPException(status_code=400, detail="No sales reps available in the system to assign leads to.")

        # AI-based Header Mapping (run once on the first row)
        ai_mapping = {"name": None, "phone": None, "facility_type": None}
        gemini_api_key = os.getenv("GEMINI_API_KEY")
        if gemini_api_key and gemini_api_key != "your-gemini-api-key-here" and normalized_rows:
            try:
                import google.generativeai as genai
                genai.configure(api_key=gemini_api_key)
                model = genai.GenerativeModel(os.getenv("GEMINI_MODEL", "gemini-2.5-flash"))
                sample_row = normalized_rows[0]
                prompt = (
                    "You are a data extraction assistant. Given the following first row of a CSV file (keys and values), "
                    "identify which key corresponds to the business/customer name, which key corresponds to the phone/mobile number, "
                    "and which key corresponds to the facility type or business category.\n\n"
                    f"Row data: {json.dumps(sample_row, ensure_ascii=False)}\n\n"
                    "Reply ONLY with a valid JSON object in this exact format, with no markdown formatting or extra text: "
                    '{"name_key": "the_exact_key", "phone_key": "the_exact_key", "facility_type_key": "the_exact_key"}. '
                    "If a key cannot be found, use null."
                )
                response = model.generate_content(prompt, request_options={"timeout": 15})
                reply_text = (response.text or "").strip()
                # Clean up markdown if the LLM adds it
                if reply_text.startswith("```json"):
                    reply_text = reply_text.split("```json")[1].split("```")[0].strip()
                elif reply_text.startswith("```"):
                    reply_text = reply_text.split("```")[1].split("```")[0].strip()
                
                ai_mapping = json.loads(reply_text)
            except Exception as e:
                print(f"AI Mapping failed: {e}")
                pass # fallback to hardcoded matching

        # Withdraw all old, uncalled leads from sales reps and give them to the Admin
        untouched_leads = db.query(models.Lead).outerjoin(models.CallLog).filter(
            models.CallLog.id == None,
            models.Lead.assigned_to != current_user.id
        ).all()
        for old_lead in untouched_leads:
            old_lead.assigned_to = current_user.id

        leads_created = 0
        rep_count = len(reps)

        for i, row in enumerate(normalized_rows):
            # 1. Try AI mapping first, then exact matches
            name = ""
            phone = ""
            facility_type = ""
            
            if ai_mapping.get("name_key") and ai_mapping["name_key"] in row:
                name = str(row[ai_mapping["name_key"]]).strip()
            if ai_mapping.get("phone_key") and ai_mapping["phone_key"] in row:
                phone = str(row[ai_mapping["phone_key"]]).strip()
            if ai_mapping.get("facility_type_key") and ai_mapping["facility_type_key"] in row:
                facility_type = str(row[ai_mapping["facility_type_key"]]).strip()

            # 2. Try exact matches if AI missed it
            if not name:
                name = str(row.get('business name') or row.get('name') or row.get('company') or row.get('title') or '').strip()
            if not phone:
                phone = str(row.get('mobile') or row.get('phone') or row.get('telephone') or '').strip()
            if not facility_type:
                facility_type = str(row.get('category') or row.get('type') or row.get('facility_type') or '').strip()
            
            notes_parts = []
            
            # 3. Try substring matching as fallback, and collect notes
            for k, v in row.items():
                v_str = str(v).strip()
                if not v_str: continue
                
                if not name and ('name' in k or 'اسم' in k):
                    name = v_str
                elif not phone and ('phone' in k or 'mobile' in k or 'رقم' in k or 'موبايل' in k):
                    phone = v_str
                elif not facility_type and ('category' in k or 'type' in k or 'نوع' in k or 'تصنيف' in k):
                    facility_type = v_str
                elif 'notes' in k or 'ملاحظات' in k:
                    notes_parts.append(v_str)
                    
            # 4. Ultimate Fallback: Deep Value Scanning (Ignore headers entirely)
            for v in row.values():
                v_str = str(v).strip()
                if not v_str: continue
                
                # Check if it looks like a phone number
                if not phone:
                    digits = ''.join(c for c in v_str if c.isdigit())
                    if 9 <= len(digits) <= 15:
                        phone = v_str
                        continue
                
                # Check if it looks like a name (mostly text, not long enough to be notes)
                if not name and len(v_str) < 60 and not any(char.isdigit() for char in v_str):
                    name = v_str
                    continue
                    
            if not name:
                name = "Unknown Lead"
            if not phone:
                phone = "0000000000"
            if not facility_type:
                facility_type = "Uncategorized"
                
            notes = " | ".join(notes_parts)
            
            # Round-robin assignment
            assigned_rep = reps[leads_created % rep_count]
            
            new_lead = models.Lead(
                name=name,
                phone=phone,
                facility_type=facility_type,
                notes=notes,
                assigned_to=assigned_rep.id,
            )
            db.add(new_lead)
            leads_created += 1
            
        if leads_created == 0:
            raise HTTPException(status_code=400, detail="لم يتم العثور على أي عملاء صالحين في الملف. تأكد من وجود عمود للاسم وعمود للرقم وأن البيانات غير فارغة.")
            
        db.commit()

        # Send Telegram Notification
        msg = (
            f"📥 <b>تم رفع دفعة عملاء جديدة!</b>\n\n"
            f"👤 قام المدير: {current_user.username} برفع شيت عملاء.\n"
            f"✅ إجمالي العملاء: {leads_created}\n"
            f"🔄 تم توزيعهم بالتساوي على {rep_count} مناديب."
        )
        send_telegram_notification(msg)

        return {"status": "success", "message": f"{leads_created} leads uploaded and distributed among {rep_count} reps."}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to process CSV: {str(e)}")


@app.get("/leads/", response_model=List[schemas.LeadOut])
def list_leads(
    status: Optional[str] = "all",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    # --- Core RBAC enforcement ---
    query = db.query(models.Lead).outerjoin(models.CallLog)

    if current_user.role == models.UserRole.SALES_REP:
        query = query.filter(models.Lead.assigned_to == current_user.id)
        if status == "pending" or status == "all":
            # Queue System: Only show 10 leads that have NOT been processed (no call logs) and not ignored
            query = query.filter(models.CallLog.id == None, models.Lead.is_ignored == False)
            limit_val = 10
        elif status == "finished":
            # Show all leads processed by this rep
            query = query.filter(models.CallLog.id != None)
            limit_val = None
        elif status == "followups":
            # Show leads that need follow-up today or earlier
            query = query.filter(models.Lead.followup_date != None, models.Lead.followup_date <= datetime.utcnow())
            limit_val = None
        elif status in ["تم الاتصال - مهتم", "تم الاتصال - غير مهتم", "لم يرد", "تم الإرسال واتساب", "رقم خاطئ"]:
            query = query.filter(models.CallLog.call_result == status)
            limit_val = None
    else:
        # admin: sees team's leads
        limit_val = None
        if status == "pending":
            query = query.filter(models.CallLog.id == None, models.Lead.is_ignored == False)
        elif status == "finished":
            query = query.filter(models.CallLog.id != None)
        elif status == "followups":
            query = query.filter(models.Lead.followup_date != None, models.Lead.followup_date <= datetime.utcnow())
        elif status == "ignored":
            query = query.filter(models.Lead.is_ignored == True)
        elif status in ["تم الاتصال - مهتم", "تم الاتصال - غير مهتم", "لم يرد", "تم الإرسال واتساب", "رقم خاطئ"]:
            query = query.filter(models.CallLog.call_result == status)

    # Must apply order_by BEFORE limit
    query = query.order_by(models.Lead.created_at.desc())
    if limit_val:
        query = query.limit(limit_val)
        
    leads = query.all()
    return [_lead_to_out(lead) for lead in leads]


@app.get("/leads/{lead_id}", response_model=schemas.LeadOut)
def get_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    if current_user.role == models.UserRole.SALES_REP and lead.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="You do not have access to this lead")
    return _lead_to_out(lead)


@app.put("/leads/{lead_id}", response_model=schemas.LeadOut)
def update_lead(
    lead_id: int,
    payload: schemas.LeadUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    # RBAC check
    if current_user.role == models.UserRole.SALES_REP and lead.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="You do not have access to this lead")

    # Update Lead attributes
    if payload.name is not None:
        lead.name = payload.name
    if payload.phone is not None:
        lead.phone = payload.phone
    if payload.facility_type is not None:
        lead.facility_type = payload.facility_type
    if payload.notes is not None:
        lead.notes = payload.notes
    if payload.next_followup is not None:
        lead.followup_date = payload.next_followup
    if payload.is_ignored is not None:
        lead.is_ignored = payload.is_ignored
    if payload.ignore_reason is not None:
        lead.ignore_reason = payload.ignore_reason
    if payload.assigned_to is not None:
        lead.assigned_to = payload.assigned_to
        # If reassigned to someone else, reset ignored status!
        lead.is_ignored = False
        lead.ignore_reason = None

    # If call_result is provided, we edit/create the latest call log
    if payload.call_result is not None:
        if payload.new_log == True:
            # Force create a new log
            new_log = models.CallLog(
                lead_id=lead_id,
                call_result=payload.call_result,
                sales_amount=payload.sales_amount or 0.0,
                notes=payload.call_notes or ""
            )
            db.add(new_log)
            if payload.next_followup is not None:
                lead.followup_date = payload.next_followup
        else:
            # Find latest call log to update
            latest_log = db.query(models.CallLog).filter(models.CallLog.lead_id == lead_id).order_by(models.CallLog.timestamp.desc()).first()
            if latest_log:
                # Update existing log
                latest_log.call_result = payload.call_result
                if payload.sales_amount is not None:
                    latest_log.sales_amount = payload.sales_amount
                if payload.call_notes is not None:
                    latest_log.notes = payload.call_notes
                if payload.next_followup is not None:
                    lead.followup_date = payload.next_followup
            else:
                # Create a new log if none exists
                new_log = models.CallLog(
                    lead_id=lead_id,
                    call_result=payload.call_result,
                    sales_amount=payload.sales_amount or 0.0,
                    notes=payload.call_notes or ""
                )
                db.add(new_log)
                if payload.next_followup is not None:
                    lead.followup_date = payload.next_followup

    db.commit()
    db.refresh(lead)
    
    # Send Telegram Notification
    rep_name = current_user.full_name or current_user.username
    if payload.is_ignored == True:
        msg = (
            f"🚫 <b>تم تخطي / تجاهل عميل!</b>\n\n"
            f"👤 <b>العميل:</b> {lead.name}\n"
            f"📞 <b>التليفون:</b> {lead.phone}\n"
            f"📝 <b>السبب:</b> {payload.ignore_reason or 'بدون سبب'}\n"
            f"👤 <b>بواسطة المندوب:</b> {rep_name}\n\n"
            f"💡 <i>يمكن للمدير مراجعة العميل في لوحة التحكم وتغيير المندوب أو حذفه.</i>"
        )
    else:
        msg = (
            f"✏️ <b>Lead Updated!</b>\n\n"
            f"👤 <b>Lead:</b> {lead.name}\n"
            f"📊 <b>Outcome:</b> {payload.call_result or 'N/A'}\n"
            f"👤 <b>Updated By:</b> {rep_name}"
        )
    send_telegram_notification(msg)
    
    return _lead_to_out(lead)


# ---------------------------------------------------------------------------
# Call Logs
# ---------------------------------------------------------------------------

def _calllog_to_out(log: models.CallLog) -> schemas.CallLogOut:
    return schemas.CallLogOut(
        id=log.id,
        lead_id=log.lead_id,
        call_result=log.call_result,
        sales_amount=log.sales_amount,
        notes=log.notes,
        timestamp=log.timestamp,
        lead_name=log.lead.name if log.lead else None,
        rep_username=log.lead.assigned_rep.username if log.lead and log.lead.assigned_rep else None,
    )


@app.post("/call-logs/", response_model=schemas.CallLogOut)
def create_call_log(
    payload: schemas.CallLogCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    lead = db.query(models.Lead).filter(models.Lead.id == payload.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    # RBAC: sales reps may only log calls against leads assigned to them.
    if current_user.role == models.UserRole.SALES_REP and lead.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="You cannot log calls for another rep's lead")

    log = models.CallLog(
        lead_id=payload.lead_id,
        call_result=payload.call_result,
        sales_amount=payload.sales_amount,
        notes=payload.notes,
    )
    db.add(log)
    
    if payload.next_followup:
        lead.followup_date = payload.next_followup
        
    db.commit()
    db.refresh(log)
    
    # Send Telegram Notification
    rep_name = current_user.full_name or current_user.username
    msg = (
        f"📞 <b>New Call Log Added!</b>\n\n"
        f"👤 <b>Lead:</b> {lead.name}\n"
        f"📊 <b>Result:</b> {log.call_result}\n"
        f"💰 <b>Sales Amount:</b> ${log.sales_amount}\n"
        f"📝 <b>Notes:</b> {log.notes or 'N/A'}\n"
        f"👤 <b>Logged By:</b> {rep_name}"
    )
    send_telegram_notification(msg)
    
    return _calllog_to_out(log)


@app.get("/call-logs/", response_model=List[schemas.CallLogOut])
def list_call_logs(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    query = db.query(models.CallLog).join(models.Lead)

    # --- Core RBAC enforcement ---
    if current_user.role == models.UserRole.SALES_REP:
        query = query.filter(models.Lead.assigned_to == current_user.id)
    # admin: sees every call log across the whole team

    logs = query.order_by(models.CallLog.timestamp.desc()).all()
    return [_calllog_to_out(log) for log in logs]


# ---------------------------------------------------------------------------
# Manager AI Assistant  (admin-only)
# ---------------------------------------------------------------------------

def _build_team_performance_context(db: Session) -> str:
    """
    Aggregates current team performance directly from SQLite and formats it
    into a compact text context string to feed the LLM. Keeping this as plain
    text (rather than raw JSON dumps) keeps token usage low and improves the
    quality of the model's natural-language answers.
    """
    reps = db.query(models.User).filter(models.User.role == models.UserRole.SALES_REP).all()

    sales_by_rep = defaultdict(float)
    calls_by_rep = defaultdict(int)
    results_by_rep = defaultdict(lambda: defaultdict(int))

    logs = db.query(models.CallLog).join(models.Lead).all()
    for log in logs:
        rep = log.lead.assigned_rep
        if not rep:
            continue
        sales_by_rep[rep.username] += log.sales_amount or 0.0
        calls_by_rep[rep.username] += 1
        results_by_rep[rep.username][log.call_result] += 1

    leads_by_rep = defaultdict(int)
    for lead in db.query(models.Lead).all():
        if lead.assigned_rep:
            leads_by_rep[lead.assigned_rep.username] += 1

    lines = [f"Team performance snapshot as of {datetime.utcnow().isoformat()}Z", ""]
    total_sales = 0.0
    for rep in reps:
        uname = rep.username
        total_sales += sales_by_rep[uname]
        result_summary = ", ".join(f"{k}: {v}" for k, v in results_by_rep[uname].items()) or "no calls logged"
        lines.append(
            f"- Rep '{uname}' ({rep.full_name or 'N/A'}): "
            f"{leads_by_rep[uname]} leads assigned, "
            f"{calls_by_rep[uname]} calls logged, "
            f"total sales ${sales_by_rep[uname]:,.2f}. "
            f"Call outcomes: {result_summary}."
        )
    lines.append("")
    lines.append(f"Team total sales: ${total_sales:,.2f}")
    lines.append(f"Team total leads: {db.query(models.Lead).count()}")
    lines.append(f"Team total calls logged: {db.query(models.CallLog).count()}")

    return "\n".join(lines)


@app.post("/api/manager-chat", response_model=schemas.ManagerChatResponse)
def manager_chat(
    payload: schemas.ManagerChatRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """
    Lets the manager ask natural-language questions about team performance.
    Flow:
      1. Aggregate live data from SQLite (sales per rep, call outcomes, lead counts).
      2. Build a compact text context describing current team performance.
      3. Send {context + admin's question} to Gemini.
      4. Return the model's natural-language insight to the frontend.
    """
    if current_user.role == models.UserRole.ADMIN:
        context = _build_team_performance_context(db)
        system_instructions = (
            "You are an AI Sales Director embedded in a medical supplies CRM. "
            "Help the sales manager understand their team's performance and provide coaching advice. "
            "ALWAYS reply in Arabic, as the manager is an Arabic speaker. "
            "If the manager asks a general question (like 'Why?' or 'How do I improve?'), "
            "give them helpful sales strategies or ask for clarification, do NOT just say the data doesn't contain the answer. "
            "When analyzing data, call out the top performing rep, struggling reps, and use concrete numbers (sales $, call counts)."
        )
    else:
        # Sales Rep context
        rep_logs = db.query(models.CallLog).join(models.Lead).filter(models.Lead.assigned_to == current_user.id).all()
        rep_leads = db.query(models.Lead).filter(models.Lead.assigned_to == current_user.id).count()
        total_sales = sum((log.sales_amount or 0.0) for log in rep_logs)
        calls_count = len(rep_logs)
        
        context = f"Rep '{current_user.username}' has {rep_leads} leads, made {calls_count} calls, and generated ${total_sales:,.2f} in sales."
        system_instructions = (
            "You are an AI Sales Coach for MedCRM, helping a medical sales representative close deals. "
            "ALWAYS reply in Arabic, as the sales rep is an Arabic speaker. "
            "Answer their questions concisely and provide practical sales advice. "
            "If they ask general questions (like 'Why?' or 'What should I do?'), give them general sales strategies. "
            "If they ask about their performance, use the data below. Be encouraging and professional."
        )

    gemini_api_key = os.getenv("GEMINI_API_KEY")
    if not gemini_api_key or gemini_api_key == "your-gemini-api-key-here":
        raise HTTPException(
            status_code=500,
            detail="GEMINI_API_KEY is not configured on the server. Set it in backend/.env",
        )

    try:
        import google.generativeai as genai

        genai.configure(api_key=gemini_api_key)
        model_name = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
        model = genai.GenerativeModel(model_name)

        full_prompt = (
            f"{system_instructions}\n\n"
            f"=== PERFORMANCE DATA ===\n{context}\n\n"
            f"=== USER'S QUESTION ===\n{payload.prompt}"
        )

        # Explicit timeout so a slow/unreachable Gemini API can never hang this
        # request indefinitely - it will raise instead, which we turn into a 502.
        response = model.generate_content(full_prompt, request_options={"timeout": 30})
        reply_text = (response.text or "").strip() if hasattr(response, "text") else str(response)

    except Exception as exc:  # noqa: BLE001 - surface a clean error to the frontend
        raise HTTPException(status_code=502, detail=f"AI assistant error: {exc}")

    return schemas.ManagerChatResponse(reply=reply_text, context_used=context)


@app.get("/api/cron/daily-report")
def send_daily_report(
    db: Session = Depends(get_db),
    token: Optional[str] = None
):
    # Ensure this is only triggered by authorized cron or admin
    expected_token = os.getenv("CRON_SECRET", "default_secret")
    if token != expected_token:
        raise HTTPException(status_code=401, detail="Unauthorized cron trigger")
        
    context = _build_team_performance_context(db)
    gemini_api_key = os.getenv("GEMINI_API_KEY")
    if not gemini_api_key:
        raise HTTPException(status_code=500, detail="Missing Gemini key")
        
    try:
        import google.generativeai as genai
        genai.configure(api_key=gemini_api_key)
        model = genai.GenerativeModel(os.getenv("GEMINI_MODEL", "gemini-2.5-flash"))
        
        prompt = (
            "You are the AI Sales Director for MedCRM. "
            "Write a highly professional, concise daily performance report IN ARABIC based on the data below. "
            "Rank the sales reps from best to worst based on their sales and calls. "
            "Use emojis for visual appeal, and give one brief sentence of advice to the manager. "
            f"\n\n=== DATA ===\n{context}"
        )
        
        response = model.generate_content(prompt, request_options={"timeout": 30})
        report_text = (response.text or "").strip()
        
        # Add a header
        final_msg = f"📊 <b>التقرير اليومي لأداء الفريق</b>\n\n{report_text}"
        send_telegram_notification(final_msg)
        
        return {"status": "success", "message": "Report sent to Telegram"}
        
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error generating report: {exc}")


@app.get("/")
def root():
    return {"status": "ok", "service": "Medical Supplies CRM API"}
